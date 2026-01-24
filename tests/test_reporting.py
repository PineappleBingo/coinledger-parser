import unittest
import pandas as pd
import os
from openpyxl import load_workbook
from src.reporting.excel_export import generate_reconciliation_report

class TestReporting(unittest.TestCase):
    def setUp(self):
        self.output_file = "test_report.xlsx"
        
        # Dummy data
        self.matched = pd.DataFrame([{
            'source_a': {'timestamp': '2025-01-01', 'asset': 'BTC', 'amount': 1.0, 'tx_id': '0x1'},
            'source_b': {'timestamp': '2025-01-01', 'asset': 'BTC', 'amount': 1.0, 'tx_id': '0x1'},
            'confidence': 1.0,
            'match_type': 'EXACT_TXID'
        }])
        
        self.conflicts = pd.DataFrame([{
            'source_a': {'timestamp': '2025-01-02', 'asset': 'ETH', 'amount': 2.0, 'tx_id': '0x2'},
            'source_b': None,
            'issue': 'MISSING_IN_BLOCKCHAIN'
        }])
        
        self.missing_in_b = pd.DataFrame([{
            'timestamp': '2025-01-03', 'asset': 'SOL', 'amount': 100.0, 'tx_id': '0x3'
        }])
        
        self.anomalies = [{
            'type': 'FEE_ANOMALY',
            'severity': 'HIGH',
            'message_kr': 'Test Anomaly',
            'tx_id': '0x1'
        }]

    def tearDown(self):
        if os.path.exists(self.output_file):
            os.remove(self.output_file)

    def test_generate_report(self):
        generate_reconciliation_report(
            self.matched, 
            self.conflicts, 
            self.missing_in_b, 
            self.anomalies, 
            self.output_file
        )
        
        self.assertTrue(os.path.exists(self.output_file))
        
        wb = load_workbook(self.output_file)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Reconciliation", wb.sheetnames)
        self.assertIn("Anomalies", wb.sheetnames)
        
        ws_recon = wb["Reconciliation"]
        # Header + 1 matched + 1 conflict + 1 missing = 4 rows
        self.assertEqual(ws_recon.max_row, 4)
        
        ws_anom = wb["Anomalies"]
        # Header + 1 anomaly = 2 rows
        self.assertEqual(ws_anom.max_row, 2)

if __name__ == '__main__':
    unittest.main()
