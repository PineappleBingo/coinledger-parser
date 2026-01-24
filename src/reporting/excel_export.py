import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict

def generate_reconciliation_report(
    matched: pd.DataFrame, 
    conflicts: pd.DataFrame, 
    missing_in_b: pd.DataFrame, 
    anomalies: List[Dict],
    output_path: str
):
    """
    Generates an Excel report with 3 sheets: Summary, Reconciliation, Anomalies.
    """
    wb = Workbook()
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    stats = [
        ("Total Matched", len(matched)),
        ("Total Conflicts", len(conflicts)),
        ("Missing in Blockchain", len(missing_in_b)),
        ("Anomalies Detected", len(anomalies))
    ]
    
    ws_summary.append(["Metric", "Count"])
    for metric, count in stats:
        ws_summary.append([metric, count])
        
    # Style Summary
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        
    # --- Sheet 2: Reconciliation ---
    ws_recon = wb.create_sheet("Reconciliation")
    
    # Headers
    headers = [
        "Status", "Confidence", "Match Type",
        "Source A (Date)", "Source A (Asset)", "Source A (Amount)", "Source A (TxID)",
        "Source B (Date)", "Source B (Asset)", "Source B (Amount)", "Source B (TxID)",
        "Issue / Deviation"
    ]
    ws_recon.append(headers)
    
    # Fills
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Light Red
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light Yellow
    
    # 1. Add Matched
    if not matched.empty:
        for _, row in matched.iterrows():
            src_a = row['source_a']
            src_b = row['source_b']
            
            row_data = [
                "MATCHED",
                f"{row['confidence']*100:.1f}%",
                row['match_type'],
                src_a.get('timestamp'), src_a.get('asset'), src_a.get('amount'), src_a.get('tx_id'),
                src_b.get('timestamp'), src_b.get('asset'), src_b.get('amount'), src_b.get('tx_id'),
                ""
            ]
            ws_recon.append(row_data)
            # Color row green
            for cell in ws_recon[ws_recon.max_row]:
                cell.fill = green_fill

    # 2. Add Conflicts
    if not conflicts.empty:
        for _, row in conflicts.iterrows():
            src_a = row['source_a']
            # source_b might be None if missing
            src_b = row.get('source_b') or {}
            
            status = "CONFLICT" if src_b else "MISSING_IN_BLOCKCHAIN"
            fill = red_fill if src_b else yellow_fill
            
            row_data = [
                status,
                f"{row.get('confidence', 0)*100:.1f}%" if 'confidence' in row else "N/A",
                row.get('match_type', ''),
                src_a.get('timestamp'), src_a.get('asset'), src_a.get('amount'), src_a.get('tx_id'),
                src_b.get('timestamp', ''), src_b.get('asset', ''), src_b.get('amount', ''), src_b.get('tx_id', ''),
                row.get('issue', '')
            ]
            ws_recon.append(row_data)
            for cell in ws_recon[ws_recon.max_row]:
                cell.fill = fill

    # 3. Add Missing in A (Present in B)
    if not missing_in_b.empty:
        for _, row in missing_in_b.iterrows():
            row_data = [
                "MISSING_IN_CEX",
                "N/A",
                "",
                "", "", "", "",
                row.get('timestamp'), row.get('asset'), row.get('amount'), row.get('tx_id'),
                "Found in Blockchain but not in CEX export"
            ]
            ws_recon.append(row_data)
            for cell in ws_recon[ws_recon.max_row]:
                cell.fill = yellow_fill

    # --- Sheet 3: Anomalies ---
    ws_anom = wb.create_sheet("Anomalies")
    ws_anom.append(["Type", "Severity", "Message (KR)", "TxID"])
    
    for anom in anomalies:
        ws_anom.append([
            anom.get('type'),
            anom.get('severity'),
            anom.get('message_kr'),
            anom.get('tx_id')
        ])
        
    # Save
    wb.save(output_path)
    print(f"Report saved to {output_path}")
